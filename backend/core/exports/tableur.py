"""
Génération de classeurs Excel — équivalent openpyxl de SpreadsheetExportService.

La mise en forme est reprise à l'identique (bandeau d'en-tête bleu, bordures
fines, volets figés, largeurs ajustées) pour que les fichiers produits restent
reconnaissables par les utilisateurs.
"""

import io
import re
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COULEUR_ENTETE = "4472C4"
COULEUR_TOTAUX = "E7ECF4"

POLICE = "Calibri"
TAILLE = 11

_BORDURE_FINE = Side(style="thin")
BORDURES = Border(
    left=_BORDURE_FINE, right=_BORDURE_FINE, top=_BORDURE_FINE, bottom=_BORDURE_FINE
)

#: Excel refuse ces caractères dans un nom d'onglet, et le limite à 31 signes.
_CARACTERES_INTERDITS = re.compile(r"[\\/*?\[\]:]")


def nom_onglet(titre):
    """Nom d'onglet valide — portage de sanitizeSheetTitle()."""
    nettoye = _CARACTERES_INTERDITS.sub(" ", str(titre or ""))[:31].strip()
    return nettoye or "Feuil1"


def _ajuster_largeurs(feuille, nombre_colonnes):
    """
    Ajuste la largeur des colonnes au contenu.

    openpyxl n'a pas d'équivalent à `setAutoSize(true)` : Excel n'applique
    l'ajustement automatique qu'à l'affichage. On calcule donc les largeurs.
    """
    for index in range(1, nombre_colonnes + 1):
        lettre = get_column_letter(index)
        largeur = max(
            (len(str(cellule.value)) for cellule in feuille[lettre] if cellule.value),
            default=0,
        )
        feuille.column_dimensions[lettre].width = min(max(largeur + 2, 10), 60)


def _styler_entete(feuille, ligne, nombre_colonnes):
    for index in range(1, nombre_colonnes + 1):
        cellule = feuille.cell(row=ligne, column=index)
        cellule.font = Font(bold=True, color="FFFFFF", size=TAILLE, name=POLICE)
        cellule.fill = PatternFill("solid", start_color=COULEUR_ENTETE)
        cellule.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )


def _border_plage(feuille, premiere_ligne, derniere_ligne, nombre_colonnes):
    for ligne in range(premiere_ligne, derniere_ligne + 1):
        for colonne in range(1, nombre_colonnes + 1):
            feuille.cell(row=ligne, column=colonne).border = BORDURES


def _ecrire_lignes(feuille, lignes, depart, nombre_colonnes):
    """Écrit les données en complétant ou tronquant chaque ligne au bon format."""
    numero = depart
    for valeurs in lignes:
        valeurs = list(valeurs)
        valeurs += [""] * (nombre_colonnes - len(valeurs))
        for index, valeur in enumerate(valeurs[:nombre_colonnes], start=1):
            feuille.cell(row=numero, column=index, value=valeur)
        numero += 1
    return numero - 1


def remplir_feuille(feuille, entetes, lignes):
    """Tableau simple : une ligne d'en-tête puis les données."""
    if not entetes:
        return

    nombre_colonnes = len(entetes)
    for index, entete in enumerate(entetes, start=1):
        feuille.cell(row=1, column=index, value=entete)

    derniere = max(1, _ecrire_lignes(feuille, lignes, 2, nombre_colonnes))

    _border_plage(feuille, 1, derniere, nombre_colonnes)
    _styler_entete(feuille, 1, nombre_colonnes)
    feuille.freeze_panes = "A2"

    if derniere > 1:
        for ligne in range(2, derniere + 1):
            for colonne in range(1, nombre_colonnes + 1):
                cellule = feuille.cell(row=ligne, column=colonne)
                cellule.font = Font(name=POLICE, size=TAILLE)
                cellule.alignment = Alignment(vertical="top")

    _ajuster_largeurs(feuille, nombre_colonnes)


def remplir_tableau_structure(
    feuille, titre_document, lignes_meta, entetes, lignes, ligne_totaux=None
):
    """Tableau avec bandeau de titre, métadonnées et ligne de totaux optionnelle."""
    if not entetes:
        return

    nombre_colonnes = len(entetes)
    derniere_lettre = get_column_letter(nombre_colonnes)
    numero = 1

    feuille.cell(row=numero, column=1, value=titre_document)
    feuille.merge_cells(f"A{numero}:{derniere_lettre}{numero}")
    titre = feuille.cell(row=numero, column=1)
    titre.font = Font(bold=True, size=14, name=POLICE, color="1F2937")
    titre.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    feuille.row_dimensions[numero].height = 24
    numero += 1

    for ligne_meta in lignes_meta or []:
        feuille.cell(row=numero, column=1, value=ligne_meta)
        feuille.merge_cells(f"A{numero}:{derniere_lettre}{numero}")
        cellule = feuille.cell(row=numero, column=1)
        cellule.font = Font(size=TAILLE, name=POLICE, color="4B5563")
        cellule.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        numero += 1

    numero += 1  # ligne vide de séparation
    ligne_entete = numero
    for index, entete in enumerate(entetes, start=1):
        feuille.cell(row=ligne_entete, column=index, value=entete)

    derniere = max(
        ligne_entete, _ecrire_lignes(feuille, lignes, ligne_entete + 1, nombre_colonnes)
    )

    if ligne_totaux:
        ligne_total = derniere + 1
        _ecrire_lignes(feuille, [ligne_totaux], ligne_total, nombre_colonnes)
        for colonne in range(1, nombre_colonnes + 1):
            cellule = feuille.cell(row=ligne_total, column=colonne)
            cellule.font = Font(bold=True, size=TAILLE, name=POLICE, color="111827")
            cellule.fill = PatternFill("solid", start_color=COULEUR_TOTAUX)
            cellule.alignment = Alignment(vertical="center")
        derniere = ligne_total

    _border_plage(feuille, ligne_entete, derniere, nombre_colonnes)
    _styler_entete(feuille, ligne_entete, nombre_colonnes)

    fin_corps = derniere - 1 if ligne_totaux else derniere
    for ligne in range(ligne_entete + 1, fin_corps + 1):
        for colonne in range(1, nombre_colonnes + 1):
            cellule = feuille.cell(row=ligne, column=colonne)
            cellule.font = Font(name=POLICE, size=TAILLE)
            cellule.alignment = Alignment(vertical="top")

    feuille.freeze_panes = f"A{ligne_entete + 1}"
    _ajuster_largeurs(feuille, nombre_colonnes)


def classeur_multi_feuilles(definitions):
    """
    Construit un classeur à partir d'une liste de définitions d'onglets.

    Chaque définition porte au minimum `titre`, `entetes` et `lignes` ; la
    présence de `titre_document` bascule sur la mise en page structurée.
    """
    classeur = Workbook()
    premier = True

    for definition in definitions:
        if premier:
            feuille = classeur.active
            premier = False
        else:
            feuille = classeur.create_sheet()
        feuille.title = nom_onglet(definition["titre"])

        if definition.get("titre_document"):
            remplir_tableau_structure(
                feuille,
                definition["titre_document"],
                definition.get("lignes_meta", []),
                definition["entetes"],
                definition["lignes"],
                definition.get("ligne_totaux"),
            )
        else:
            remplir_feuille(feuille, definition["entetes"], definition["lignes"])

    classeur.active = 0
    return classeur


def reponse_xlsx(classeur, nom_fichier):
    """Réponse HTTP de téléchargement du classeur."""
    if not nom_fichier.lower().endswith(".xlsx"):
        nom_fichier += ".xlsx"

    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)

    reponse = HttpResponse(
        tampon.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    reponse["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
    reponse["Cache-Control"] = "max-age=0"
    return reponse


def horodatage():
    """Suffixe de nom de fichier, au format employé par Laravel."""
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")


def classeur_simple(titre, entetes, lignes):
    """Raccourci : un classeur d'un seul onglet."""
    return classeur_multi_feuilles(
        [{"titre": titre, "entetes": entetes, "lignes": lignes}]
    )
